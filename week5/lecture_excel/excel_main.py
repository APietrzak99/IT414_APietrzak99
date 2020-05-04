import openpyxl

my_workbook = openpyxl.Workbook()
curr_sheet = my_workbook.active
curr_sheet.title = "My Sheet"
my_workbook.create_sheet(title = "Another Sheet")
curr_sheet = my_workbook["Another Sheet"]
curr_sheet["B3"] = "My Cell"

birthdays = [{"name": "Chris" ,"birthday": "1/23/1975"},
{"name": "Diane","birthday": "2/23/1982"},
{"name": "Anthony","birthday":"11/18/1999"},
{"name": "Jamal","birthday":"11/07/1987"}]

curr_sheet = my_workbook["My Sheet"]
row_count = 1

for item in birthdays:
    col_count = 1
    curr_sheet.cell(row = row_count, column = col_count).value = item["name"]
    col_count = col_count + 1
    curr_sheet.cell(row = row_count, column = col_count).value = item["birthday"]
    row_count = row_count + 1

curr_sheet.column_dimensions["B"].width = 50


my_workbook.save("files/my_workbook.xlsx")