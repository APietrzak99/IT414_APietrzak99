import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import *
import ezsheets
import requests
from bs4 import BeautifulSoup
import collections
import webbrowser

my_url = "https://ool-content.walshcollege.edu/CourseFiles/IT/IT414/MASTER/Week05/WI20-Assignment/sales_data.html"
request = requests.get(my_url)

if request.status_code == 200:
    sales_data = open("files/sales_data.html", "wb")
    for data_chunk in request.iter_content(100000):
        sales_data.write(data_chunk)

my_sales_page = open("files/sales_data.html")

sales_soup = BeautifulSoup(my_sales_page, "lxml")

data = [[cell.get_text(strip=True) for cell in row.find_all('td')[3:5]]
        for row in sales_soup.find_all("tr")]

temp_list = []

for val in data[1:]:
    temp_list.append({"Product":val[0], "Quantity": val[1]})

result = collections.Counter()

for data in temp_list:
    result[data['Product']] += int(data['Quantity'])

combined_quantities = [{product: str(quantity)} for product, quantity in result.items()]

sales_workbook = openpyxl.Workbook()
curr_sheet = sales_workbook.active
curr_sheet.title = "Data Sheet"
curr_sheet.merge_cells("A1:E1")
curr_sheet["A1"] = "Sales Data"
curr_sheet["A2"] = "Product"
curr_sheet["B2"] = "Quantity"
curr_sheet["A2"].font = Font(b=True)
curr_sheet["B2"].font = Font(b=True)

row_count = 3
for val in combined_quantities:
    if isinstance(val,dict):
       for key, value in val.items():
            curr_sheet.cell(row=row_count, column=1).value = str(key)
            curr_sheet.cell(row=row_count, column=2).value = int(value)
    row_count = row_count + 1

curr_sheet["A1"].font = Font(sz=20.0, b=True, color = "33E8FF")
curr_sheet["A1"].alignment = Alignment(horizontal="center")

max_row = curr_sheet.max_row
max_col = curr_sheet.max_column

new_max_row = max_row + 1
avg_col_start = 2

curr_sheet["A" + str(new_max_row)].font = Font(b=True)
curr_sheet["A" + str(new_max_row)].alignment = Alignment(wrap_text=True, horizontal="center")
curr_sheet["A" + str(new_max_row)].value = "Average:"

sales_workbook.save("files/sales_data.xlsx")

row_num = 3

curr_sheet.cell(row=new_max_row, column=avg_col_start).value = "=ROUND((AVERAGE(B" + str(row_num) + ":" + "B" + str(new_max_row - 1) + ")), 2)"

sales_workbook.save("files/sales_data.xlsx")

my_spreadsheet = ezsheets.createSpreadsheet("Sales Data")
my_url = my_spreadsheet.url

curr_sheet = my_spreadsheet[0]
rows = curr_sheet.getRows()

rows[0][0] = "Product"
rows[0][1] = "Quantity"

row_count_goog = 1
for val in combined_quantities:
    if isinstance(val,dict):
       for key, value in val.items():
            rows[row_count_goog][0] = str(key)
            rows[row_count_goog][1] = int(value)
    row_count_goog = row_count_goog + 1

curr_sheet.updateRows(rows)

webbrowser.open(my_url)